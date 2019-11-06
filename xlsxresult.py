#xslx2mlf
# coding:utf-8
from openpyxl import *
from openpyxl.styles import Font,colors,Color
import re
font=Font(color=colors.RED)
wb=load_workbook('result.xlsx')
ws=wb['ditu']
fw=open('wanmoerji_skill.txt','r')
lines=fw.readlines()
i=1
#print(lines[0])
while(1):
    if(ws.cell(row=1+i,column=2).value):
        if isinstance(ws.cell(row=1+i,column=8).value,long):
            line=ws.cell(row=1+i,column=8).value
	    #print(line)
            skill=lines[line-1]
	    
            index=skill.find('>')
	    #print(skill[index+1:])
            ws.cell(row=1+i,column=3,value=skill[index+1:])
        elif isinstance(ws.cell(row=1+i,column=9).value,long):
            line=int(ws.cell(row=1+i,column=9).value)
            skill=lines[line-1]
            index=skill.find('>')
            ws.cell(row=1+i,column=3,value=skill[index+1:])
        elif isinstance(ws.cell(row=1+i,column=10).value,long):
            line=int(ws.cell(row=1+i,column=10).value)
	    #print(line)
	    #print(ws.cell(row=1+i,column=2).value)
            skill=lines[line-1]
            index=skill.find('>')
            #print(skill[index+1:])
            ws.cell(row=1+i,column=3,value=skill[index+1:])
        else:
            ws.cell(row=1+i,column=3,value=0)
        if ws.cell(row=1+i,column=3).value[:-1] != ws.cell(row=1+i,column=4).value:
            ws.cell(row=1+i,column=4).font=font
        i+=1
    else:break

wb.save('newtestresult.xlsx')
fw.close()    
